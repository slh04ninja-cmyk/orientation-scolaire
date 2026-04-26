import os
from io import StringIO
import pandas as pd
import numpy as np
import streamlit as st

# ------------------------------
# Mapping arabe → français (matières MASSAR)
# ------------------------------
ARABIC_TO_FR = {
    "الرياضيات": "Mathematiques",
    "علوم الحياة والأرض": "SVT",
    "الفيزياء والكيمياء": "Physique",
    "الفرنسية": "Francais",
    "اللغة العربية": "Arabe",
    "الفلسفة": "Philosophie",
    "التربية الإسلامية": "Islam",
}

SUBJECT_PATTERNS = {
    "Mathematiques": ["mathematiques", "maths", "math", "mathe"],
    "Physique": ["physique", "phys", "chimie"],
    "SVT": ["svt", "biologie", "sciences de la vie"],
    "Francais": ["francais", "lf", "langue francaise"],
}

TRACK_ORDER = ["SM", "SE", "LT"]
TRACK_NAMES = {
    "SM": "Sciences Mathematiques",
    "SE": "Sciences Experimentales",
    "LT": "Lettres et Traduction",
}
TRACK_COLORS = {"SM": "#7c6cf0", "SE": "#22c997", "LT": "#f06292"}

DEFAULT_WEIGHTS = {
    "Mathematiques": 3.0,
    "Physique": 3.0,
    "SVT": 2.0,
    "Francais": 2.0,
}

DEFAULT_THRESHOLDS = {"SM": 70, "SE": 50, "LT": 30}

# ✅ Mapping matières arabes vers abréviations français pour tableau
SUBJECT_ABBREV = {
    "الرياضيات": "moy. Math",
    "الفيزياء والكيمياء": "moy. PC",
    "الفيزياء": "moy. PC",
    "الكيمياء": "moy. PC",
    "علوم الحياة والأرض": "moy. SVT",
    "علوم الحياة": "moy. SVT",
    "اللغة الفرنسية": "moy. Fr",
    "الفرنسية": "moy. Fr",
    "الفرنسيـة": "moy. Fr",
    "Mathematiques": "moy. Math",
    "Physique": "moy. PC",
    "SVT": "moy. SVT",
    "Francais": "moy. Fr",
}


def get_css():
    return """
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'Outfit', sans-serif; }
.stApp { background: #0e0e1a; color: #e0e0e0; }

section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0b0b18 0%, #151530 100%);
    border-right: 1px solid rgba(255, 255, 255, 0.06);
}

.block-container { padding-top: 2rem; max-width: 1200px; }

.metric-card {
    background: rgba(255, 255, 255, 0.03);
    border: 1px solid rgba(255, 255, 255, 0.07);
    border-radius: 18px;
    padding: 1.6rem 1.4rem;
    backdrop-filter: blur(12px);
    transition: transform 0.3s ease, box-shadow 0.3s ease;
}
.metric-card:hover {
    transform: translateY(-3px);
    box-shadow: 0 14px 44px rgba(0, 0, 0, 0.35);
}

.track-badge {
    display: inline-block;
    padding: 5px 14px;
    border-radius: 30px;
    font-weight: 600;
    font-size: 0.8rem;
    margin: 2px;
}

[data-testid="stFileUploadDropzone"] {
    border: 2px dashed rgba(255, 255, 255, 0.12) !important;
    border-radius: 16px !important;
}
[data-testid="stDataFrame"] { border-radius: 14px; overflow: hidden; }
hr { border-color: rgba(255, 255, 255, 0.06) !important; }
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-thumb { background: rgba(255, 255, 255, 0.12); border-radius: 6px; }

.header-title {
    font-size: 2.4rem;
    margin-bottom: 0.3rem;
    background: linear-gradient(135deg, #7c6cf0, #22c997, #f06292);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.header-subtitle { color: rgba(255, 255, 255, 0.55); font-size: 1.05rem; max-width: 700px; margin: auto; }

.welcome-icon { font-size: 4rem; margin-bottom: 0.8rem; }
.welcome-title { color: rgba(255, 255, 255, 0.75); font-weight: 400; }

.score-row { margin: 8px 0; padding: 10px 14px; border-radius: 12px; }
.score-row.eligible { background: rgba(255, 255, 255, 0.04); border: 1px solid rgba(255, 255, 255, 0.1); }
.score-row.not-eligible { background: rgba(255, 255, 255, 0.015); border: 1px solid rgba(255, 255, 255, 0.04); }
.score-label { font-weight: 600; margin-left: 6px; }
.score-value { color: rgba(255, 255, 255, 0.7); font-weight: 600; font-family: 'JetBrains Mono', monospace; }
.score-threshold { font-size: 0.7rem; color: rgba(255, 255, 255, 0.35); }

.progress-bar-bg { background: rgba(255, 255, 255, 0.08); border-radius: 8px; height: 8px; }
.progress-bar-fill { height: 100%; border-radius: 8px; transition: width 0.5s ease; }

.seuil-container { margin-top: 0.8rem; padding: 0.8rem 1rem; background: rgba(255, 255, 255, 0.03); border-radius: 12px; border: 1px solid rgba(255, 255, 255, 0.06); font-size: 0.85rem; }
.seuil-track { position: relative; height: 30px; margin: 8px 0; }
.seuil-bar { position: absolute; top: 12px; left: 0; right: 0; height: 6px; background: linear-gradient(90deg, #f06292, #22c997, #7c6cf0); border-radius: 6px; }
.seuil-marker { position: absolute; top: 6px; width: 18px; height: 18px; border-radius: 50%; border: 2px solid #0e0e1a; transform: translateX(-50%); }
.seuil-labels { display: flex; justify-content: space-between; color: rgba(255, 255, 255, 0.4); font-size: 0.75rem; }

.student-card { text-align: center; padding: 2rem 1.4rem; }
.student-name { color: #fff; margin-bottom: 0.6rem; }
.student-score { font-size: 2.4rem; font-weight: 800; }
.student-score-unit { font-size: 1rem; font-weight: 400; color: rgba(255, 255, 255, 0.4); }

.color-sm { color: #7c6cf0; }
.color-se { color: #22c997; }
.color-lt { color: #f06292; }
"""


def extract_massar_metadata(uploaded_file):
    metadata = {
        "matiere": "Matière inconnue",
        "classe_complete": "Classe inconnue",
        "groupe": "Groupe inconnu"
    }
    try:
        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0)
        from openpyxl import load_workbook
        wb = load_workbook(uploaded_file)
        ws = wb.active

        classe = ws.cell(row=9, column=4).value
        if pd.notna(classe) and isinstance(classe, str):
            metadata["classe_complete"] = classe.strip()

        groupe = ws.cell(row=9, column=9).value
        if pd.notna(groupe) and isinstance(groupe, str):
            metadata["groupe"] = groupe.strip()

        matiere = ws.cell(row=11, column=15).value
        if pd.notna(matiere) and isinstance(matiere, str):
            metadata["matiere"] = matiere.strip()
    except Exception:
        pass
    return metadata


def read_massar_format(uploaded_file):
    try:
        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0)
        df_raw = pd.read_excel(uploaded_file, header=None, engine="openpyxl")
        if len(df_raw) < 17:
            raise ValueError("Fichier trop court pour format MASSAR")

        codes_row = df_raw.iloc[14].tolist() if len(df_raw) > 14 else []
        subheaders_row = df_raw.iloc[16].tolist() if len(df_raw) > 16 else []

        note_cols = []
        for i in range(len(codes_row)):
            try:
                code = codes_row[i]
                if pd.notna(code) and isinstance(code, str) and '#' in code:
                    if code in ['#1#', '#2#', '#3#']:
                        if i < len(subheaders_row):
                            subh = subheaders_row[i]
                            if pd.notna(subh) and subh == 'النقطة':
                                note_cols.append(i)
            except (IndexError, TypeError):
                continue

        if len(note_cols) < 3:
            raise ValueError(f"Format MASSAR invalide : {len(note_cols)} devoirs trouvés")

        cols_to_keep = [1, 2, 3] + note_cols
        df = df_raw.iloc[17:, cols_to_keep].copy()
        col_names = ["ID", "Num_Massar", "Eleve"] + [f"Devoir_{i+1}" for i in range(len(note_cols))]
        df.columns = col_names

        for col in col_names[2:]:
            df[col] = pd.to_numeric(df[col], errors="coerce")

        df = df.dropna(how="all")
        df = df[df["Eleve"].notna()].copy()
        if len(df) == 0:
            # Debug: montrer ce qu'on essaie d'extraire
            sample = df_raw.iloc[17:22, cols_to_keep].to_dict('records')
            raise ValueError(
                f"Aucune donnée élève extraite. "
                f"Codes trouvés: {len(note_cols)} devoirs aux colonnes {note_cols}. "
                f"Colonnes extraites: {cols_to_keep}. "
                f"Échantillon lignes 17-21: {sample}"
            )
        return df
    except Exception as e:
        raise e


def read_excel_safe(uploaded_file, try_massar=True):
    if try_massar:
        try:
            return read_massar_format(uploaded_file)
        except Exception:
            pass

    errors = []
    for engine in ("openpyxl", "xlrd", "calamine"):
        try:
            uploaded_file.seek(0)
            return pd.read_excel(uploaded_file, engine=engine)
        except Exception as e:
            errors.append(f"{engine}: {str(e)}")

    for sep in ("\t", ";", ","):
        try:
            uploaded_file.seek(0)
            raw = uploaded_file.read().decode("utf-8", errors="replace")
            df = pd.read_csv(StringIO(raw), sep=sep)
            if len(df.columns) > 1:
                return df
        except Exception as e:
            errors.append(f"csv ({sep}): {str(e)}")

    try:
        uploaded_file.seek(0)
        raw = uploaded_file.read().decode("utf-8", errors="replace")
        df = pd.read_csv(StringIO(raw), sep=None, engine="python")
        if len(df.columns) > 1:
            return df
    except Exception as e:
        errors.append(f"csv_auto: {str(e)}")

    raise RuntimeError("Lecture impossible.\n" + "\n".join(errors))


def clean_dataframe(df):
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df = df.dropna(how="all")
    return df


def detect_subject_columns(df):
    devoir_cols = [col for col in df.columns if isinstance(col, str) and col.startswith("Devoir_")]
    if devoir_cols:
        return {"Matiere": devoir_cols}

    mapping = {}
    for col in df.columns:
        if not isinstance(col, str):
            continue
        low = col.lower().strip()
        for subject, keywords in SUBJECT_PATTERNS.items():
            if any(kw in low for kw in keywords):
                mapping.setdefault(subject, []).append(col)
                break
    return mapping


def detect_student_columns(df):
    nom = None
    prenom = None
    complet = None
    for col in df.columns:
        low = col.lower().strip()
        if any(k in low for k in ["nom et prenom", "nom_complet", "nom complet", "eleve"]):
            complet = col
        elif any(k in low for k in ["prenom", "first name"]):
            prenom = col
        elif any(k in low for k in ["nom", "name", "last name", "famille"]):
            nom = col
    return nom, prenom, complet


def build_student_name(df):
    if "Eleve" in df.columns:
        return df["Eleve"].astype(str)
    nom, prenom, complet = detect_student_columns(df)
    if complet:
        return df[complet].astype(str)
    if nom and prenom:
        return df[nom].astype(str) + " " + df[prenom].astype(str)
    if nom:
        return df[nom].astype(str)
    return pd.Series(["Eleve " + str(i + 1) for i in range(len(df))], index=df.index)


def compute_averages(df, subject_cols):
    avgs = pd.DataFrame(index=df.index)
    if not isinstance(subject_cols, dict):
        raise TypeError(f"subject_cols doit être un dictionnaire, reçu: {type(subject_cols)}")
    for subject, cols in subject_cols.items():
        if not isinstance(cols, (list, tuple)):
            raise TypeError(f"Pour matière '{subject}', colonnes liste attendue, reçu: {type(cols)}")
        cols_present = [c for c in cols if c in df.columns]
        if not cols_present:
            continue
        for c in cols_present:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        avgs[subject] = df[cols_present].mean(axis=1).round(2)
    return avgs.fillna(0)


def compute_student_score(averages, weights):
    if averages.empty:
        return pd.Series(dtype=float)
    if "Matiere" in averages.columns and len(averages.columns) == 1:
        return (averages["Matiere"] * 5).round(2)

    total_weight = sum(weights.get(s, 0) for s in averages.columns if s in weights)
    if total_weight == 0:
        return pd.Series(0.0, index=averages.index)

    score = pd.Series(0.0, index=averages.index)
    for subject in averages.columns:
        w = weights.get(subject, 0)
        if w > 0:
            score += w * averages[subject]
    return ((score / total_weight) * 5).round(2)


def classify_all(student_names, averages, scores, thresholds):
    rows = []
    for idx in range(len(averages)):
        score = scores.iloc[idx]
        row = {
            "Eleve": student_names.iloc[idx],
            "Score": score,
        }
        eligible = []
        for track in TRACK_ORDER:
            if score >= thresholds.get(track, 0):
                eligible.append(track)

        row["Filiere principale"] = eligible[0] if eligible else "---"
        row["Eligibilite"] = " | ".join(eligible) if eligible else "Aucune"
        row["Nb filieres"] = len(eligible)

        for subj in averages.columns:
            row["Moy. " + subj] = averages.iloc[idx].get(subj, 0)

        rows.append(row)

    result = pd.DataFrame(rows)
    if result.empty:
        return result
    result = result.sort_values("Score", ascending=False).reset_index(drop=True)
    result.index = result.index + 1
    result.index.name = "Rang"
    return result


def html_header():
    s = '<div style="text-align:center;padding:.6rem 0 1.8rem">'
    s += '<h1 class="header-title">Systeme d\'Orientation Scolaire</h1>'
    s += '<p class="header-subtitle">'
    s += 'Score unique par eleve, seuils et eligibilite multiple : '
    s += '<b class="color-sm">SM</b> &middot; '
    s += '<b class="color-se">SE</b> &middot; '
    s += '<b class="color-lt">LT</b>'
    s += '</p></div>'
    return s


def html_welcome():
    s = '<div style="text-align:center;padding:3.5rem 1rem">'
    s += '<div class="welcome-icon">&#128194;</div>'
    s += '<h2 class="welcome-title">Importez un fichier Excel pour demarrer</h2>'
    s += '</div>'
    return s


def html_seuil_scale(s_sm, s_se, s_lt):
    sm_pct = max(0, min(100, s_sm))
    se_pct = max(0, min(100, s_se))
    lt_pct = max(0, min(100, s_lt))
    s = '<div class="seuil-container">'
    s += '<div style="color:rgba(255,255,255,.5);margin-bottom:.5rem">Echelle des seuils</div>'
    s += '<div class="seuil-track">'
    s += '<div class="seuil-bar"></div>'
    s += '<div class="seuil-marker" style="left:' + str(lt_pct) + '%;background:#f06292"></div>'
    s += '<div class="seuil-marker" style="left:' + str(se_pct) + '%;background:#22c997"></div>'
    s += '<div class="seuil-marker" style="left:' + str(sm_pct) + '%;background:#7c6cf0"></div>'
    s += '</div>'
    s += '<div class="seuil-labels">'
    s += '<span>0</span>'
    s += '<span class="color-lt">LT &ge;' + str(s_lt) + '</span>'
    s += '<span class="color-se">SE &ge;' + str(s_se) + '</span>'
    s += '<span class="color-sm">SM &ge;' + str(s_sm) + '</span>'
    s += '<span>100</span>'
    s += '</div></div>'
    return s


def html_stat_card(count, color, subtitle, pct):
    s = '<div class="metric-card" style="border-left:4px solid ' + color + '">'
    s += '<div style="color:' + color + ';font-size:2rem;font-weight:700">' + str(count) + '</div>'
    s += '<div style="color:rgba(255,255,255,.7);font-size:.85rem">' + subtitle + '</div>'
    s += '<div style="color:rgba(255,255,255,.35);font-size:.8rem">' + pct + '</div>'
    s += '</div>'
    return s


def html_student_card(name, score, primary, pname, pcol):
    s = '<div class="metric-card student-card">'
    s += '<h3 class="student-name">' + name + '</h3>'
    s += '<div class="student-score" style="color:' + pcol + '">'
    s += f"{score:.1f}"
    s += '<span class="student-score-unit">/100</span></div>'
    s += '<div style="margin-top:.8rem">'
    s += '<span class="track-badge" style="background:' + pcol + '22;color:' + pcol
    s += ';border:2px solid ' + pcol + '">'
    s += 'Recommandation : ' + primary + ' - ' + pname
    s += '</span></div></div>'
    return s


def html_score_row(track, name, score, threshold, color, eligible):
    cls = "eligible" if eligible else "not-eligible"
    icon = "✅" if eligible else "❌"
    pct = min(score, 100)
    s = '<div class="score-row ' + cls + '">'
    s += '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px">'
    s += '<span>'
    s += '<span style="font-size:1rem">' + icon + '</span>'
    s += '<span class="score-label" style="color:' + color + '">' + track + ' - ' + name + '</span>'
    s += '</span>'
    s += '<span class="score-value">' + f"{score:.1f}"
    s += ' <span class="score-threshold">/ seuil ' + str(threshold) + '</span></span>'
    s += '</div>'
    s += '<div class="progress-bar-bg">'
    s += '<div class="progress-bar-fill" style="background:' + color + ';width:' + f"{pct}" + '%"></div>'
    s += '</div></div>'
    return s


def help_text():
    lines = []
    lines.append("**Score unique par eleve** :")
    lines.append("")
    lines.append("```")
    lines.append("Score = (p1*Math + p2*Phys + p3*SVT + p4*Fr)")
    lines.append("        / (p1 + p2 + p3 + p4)  *  5")
    lines.append("```")
    lines.append("")
    lines.append("Le `* 5` convertit la note /20 en score /100.")
    lines.append("")
    lines.append("**Eligibilite** : score compare au seuil de chaque filiere.")
    lines.append("")
    lines.append("**Filiere principale** : seuil le plus eleve atteint (SM > SE > LT).")
    lines.append("")
    lines.append("**Classement** : tries par score decroissant.")
    return "\n".join(lines)
# ───── À coller juste après la fin de la Partie 1 ─────

def get_moyennes_from_massar_df(df):
    """
    Calcule la moyenne des devoirs (colonnes Devoir_*) par élève.
    Retourne une Series indexée par le nom de l'élève.
    """
    devoir_cols = [c for c in df.columns if c.startswith("Devoir_")]
    if not devoir_cols:
        return pd.Series(dtype=float)
    # Garder aussi Num_Massar s'il existe
    cols_index = ["Eleve"]
    if "Num_Massar" in df.columns:
        return df.set_index("Eleve")[devoir_cols].mean(axis=1).round(2), df.set_index("Eleve")["Num_Massar"].first()
    return df.set_index("Eleve")[devoir_cols].mean(axis=1).round(2), None


def process_multiple_files(uploaded_files):
    """
    Traite une liste de fichiers (MASSAR ou standard) et retourne :
    - merged_df : DataFrame avec colonne 'Eleve' + une colonne de moyenne par matière
    - matieres_detectees : liste des noms de matière (français)
    - classe : chaîne représentant la classe (ou 'Inconnue')
    - devoirs_par_matiere : dict {matiere: nb total de devoirs}
    - massar_ids : dict {nom_eleve: num_massar}
    """
    matiere_series = {}          # matière -> liste de Series (index = nom élève)
    devoirs_counts = {}          # matière -> nombre cumulé de devoirs
    massar_ids = {}              # nom_eleve -> numéro MASSAR
    classe = "Inconnue"

    for uploaded_file in uploaded_files:
        # ---- Essayer le format MASSAR ----
        try:
            df_massar = read_massar_format(uploaded_file)
            # Métadonnées
            metadata = extract_massar_metadata(uploaded_file)
            matiere_raw = metadata["matiere"]
            matiere = ARABIC_TO_FR.get(matiere_raw, matiere_raw)
            if classe == "Inconnue" and metadata["groupe"] != "Groupe inconnu":
                classe = metadata["groupe"]

            result = get_moyennes_from_massar_df(df_massar)
            series_moy = result[0]
            num_massar_series = result[1]
            nb_devoirs = len([c for c in df_massar.columns if c.startswith("Devoir_")])

            matiere_series.setdefault(matiere, []).append(series_moy)
            devoirs_counts[matiere] = devoirs_counts.get(matiere, 0) + nb_devoirs

            # Collecter les numéros MASSAR
            if num_massar_series is not None:
                for nom, num in num_massar_series.items():
                    if pd.notna(num) and nom not in massar_ids:
                        massar_ids[nom] = str(num)

            continue          # format traité, passer au fichier suivant
        except Exception as e:
            st.warning(f"⚠️ Format MASSAR échoué pour {uploaded_file.name} : {e}")

        # ---- Fallback standard (sans retenter MASSAR) ----
        try:
            df = read_excel_safe(uploaded_file, try_massar=False)  # <-- modification ici
            df = clean_dataframe(df)
            subject_cols = detect_subject_columns(df)

            # Choix de la matière :
            if subject_cols:
                # Prendre la première matière trouvée
                matiere = list(subject_cols.keys())[0]
            else:
                # Utiliser le nom du fichier (sans extension) comme matière
                matiere = os.path.splitext(uploaded_file.name)[0]

            # Créer une colonne 'Eleve' pour l'indexation
            student_names = build_student_name(df)
            df_temp = df.copy()
            df_temp["Eleve"] = student_names

            # Colonnes de cette matière (ou toutes les colonnes numériques si aucune détectée)
            if matiere in subject_cols:
                cols_present = [c for c in subject_cols[matiere] if c in df_temp.columns]
            else:
                cols_present = df_temp.select_dtypes(include=np.number).columns.tolist()

            if cols_present:
                # Moyenne de l'élève sur ces colonnes
                series_moy = df_temp.set_index("Eleve")[cols_present].mean(axis=1).round(2)
            else:
                series_moy = pd.Series(dtype=float)

            matiere_series.setdefault(matiere, []).append(series_moy)
            devoirs_counts[matiere] = devoirs_counts.get(matiere, 0) + len(cols_present)

            # Récupération de la classe (optionnelle, pas toujours présente)
            if classe == "Inconnue" and "Classe" in df_temp.columns:
                classe = str(df_temp["Classe"].iloc[0]) if len(df_temp) > 0 else "Inconnue"

        except Exception as e:
            # On ignore le fichier problématique
            st.warning(f"⚠️ Impossible de traiter {uploaded_file.name} – {e}")

    # ----------------------------------------------------
    # Fusionner toutes les séries par matière
    # ----------------------------------------------------
    if not matiere_series:
        raise ValueError("Aucune donnée élève exploitable dans les fichiers fournis.")

    merged_data = {}
    for matiere, series_list in matiere_series.items():
        # Concaténer toutes les séries de cette matière (si plusieurs fichiers)
        all_series = pd.concat(series_list, axis=1)
        # Moyenne si plusieurs fichiers pour la même matière
        merged_data[matiere] = all_series.mean(axis=1).round(2)

    merged_df = pd.DataFrame(merged_data)
    merged_df.index.name = "Eleve"
    merged_df.reset_index(inplace=True)
    merged_df.fillna(0, inplace=True)

    matieres_detectees = list(merged_data.keys())
    return merged_df, matieres_detectees, classe, devoirs_counts, massar_ids
